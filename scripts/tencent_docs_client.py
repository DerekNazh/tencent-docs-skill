"""腾讯文档 Open API 客户端 — 封装所有 API 调用。"""

from __future__ import annotations

import io
import json
import tempfile
import time
from pathlib import Path
from typing import Any

import openpyxl
import requests

from tencent_docs_config import TencentDocsAuth, get_base_url


class TencentDocsError(Exception):
    """腾讯文档 API 错误。"""

    def __init__(self, ret: int, msg: str) -> None:
        self.ret = ret
        self.msg = msg
        super().__init__(f"腾讯文档 API 错误 (ret={ret}): {msg}")


# 常见错误码中文映射
_ERROR_MESSAGES: dict[int, str] = {
    400006: "Access Token 无效或已过期，请重新授权",
    400007: "需要超级会员 (VIP) 权限",
    400008: "积分不足，AI 生成配额已用完",
    11607: "请求参数错误，请检查 fileID / type 等字段",
    -32603: "请求参数错误，请检查 fileID / type 等字段",
}


class TencentDocsClient:
    """腾讯文档 Open API 客户端。"""

    def __init__(self, auth: TencentDocsAuth) -> None:
        self._auth = auth
        self._base_url = get_base_url()
        self._session = requests.Session()
        self._session.headers.update(auth.as_headers())
        # 自动跟随重定向（腾讯文档部分接口返回 301）
        self._session.max_redirects = 10

    def _url(self, path: str) -> str:
        """拼接完整 URL。path 以 / 开头，如 /drive/v2/folders/%2F/。"""
        return f"{self._base_url}{path}"

    def _check_response(self, resp: dict[str, Any]) -> dict[str, Any]:
        """检查 API 响应，ret != 0 时抛出 TencentDocsError。"""
        ret = resp.get("ret", -1)
        if ret != 0:
            msg = resp.get("msg", "未知错误")
            hint = _ERROR_MESSAGES.get(ret, "")
            if hint:
                msg = f"{msg} — {hint}"
            raise TencentDocsError(ret, msg)
        return resp.get("data", {})

    def _get(self, path: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        """GET 请求。"""
        resp = self._session.get(self._url(path), params=params, timeout=30)
        resp.raise_for_status()
        return self._check_response(resp.json())

    def _post(
        self,
        path: str,
        data: dict[str, Any] | None = None,
        json_body: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """POST 请求，支持 form-urlencoded 和 JSON body。"""
        if json_body is not None:
            resp = self._session.post(
                self._url(path), json=json_body, timeout=30
            )
        else:
            resp = self._session.post(self._url(path), data=data, timeout=30)
        resp.raise_for_status()
        return self._check_response(resp.json())

    def _patch(
        self,
        path: str,
        data: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """PATCH 请求。"""
        resp = self._session.patch(self._url(path), data=data, timeout=30)
        resp.raise_for_status()
        return self._check_response(resp.json())

    def _delete(self, path: str) -> dict[str, Any]:
        """DELETE 请求。"""
        resp = self._session.delete(self._url(path), timeout=30)
        resp.raise_for_status()
        return self._check_response(resp.json())

    # ── 验证 ──────────────────────────────────────────────

    def verify(self) -> dict[str, Any]:
        """验证认证是否有效，返回 API 用量信息。"""
        return self._get("/drive/v2/util/resource-use")

    # ── 文件夹 ────────────────────────────────────────────

    def list_folder(
        self,
        folder_id: str = "/",
        list_type: str = "folder",
        sort_type: str = "browse",
        asc: int = 0,
    ) -> list[dict[str, Any]]:
        """列出文件夹内容。

        Args:
            folder_id: 文件夹 ID，根目录为 "/"。
            list_type: 列表类型，默认 "folder"。
            sort_type: 排序方式，默认 "browse"。
            asc: 升序 1 / 降序 0。

        Returns:
            文件/文件夹列表。
        """
        from urllib.parse import quote

        encoded_id = quote(folder_id, safe="")
        # 根目录需要尾部 / 避免 301 循环
        path = f"/drive/v2/folders/{encoded_id}/"
        params = {"listType": list_type, "sortType": sort_type, "asc": asc}
        data = self._get(path, params=params)
        return data.get("list", [])

    # ── 搜索 ──────────────────────────────────────────────

    def search(self, keyword: str) -> list[dict[str, Any]]:
        """按关键词搜索文件。"""
        data = self._get("/drive/v2/search", params={"searchName": keyword})
        return data.get("list", [])

    # ── 元数据 ────────────────────────────────────────────

    def read_metadata(self, file_id: str) -> dict[str, Any]:
        """读取文件元数据。"""
        from urllib.parse import quote

        encoded_id = quote(file_id, safe="")
        return self._get(f"/drive/v2/files/{encoded_id}/metadata")

    # ── 文档内容 ──────────────────────────────────────────

    def read_doc(self, file_id: str, export_type: str = "pdf") -> str:
        """读取在线文档内容（异步导出后保存）。

        注意：腾讯文档 API 对 doc 类型仅支持 pdf 导出，对 sheet 支持 xlsx。

        Args:
            file_id: 文件 ID。
            export_type: 导出格式 (pdf / docx / xlsx)，默认 pdf。

        Returns:
            导出后的文件路径。
        """
        content = self.async_export(file_id, export_type=export_type)
        # 根据实际下载内容判断扩展名
        ext_map = {"pdf": "pdf", "docx": "docx", "xlsx": "xlsx"}
        ext = ext_map.get(export_type, "bin")
        suffix = file_id.split("$")[-1] if "$" in file_id else file_id
        output_path = f"export_{suffix}.{ext}"
        Path(output_path).write_bytes(content)
        return output_path

    # ── 表格 ──────────────────────────────────────────────

    def read_sheet(
        self, file_id: str, sheet_id: str, cell_range: str = "A1:Z100"
    ) -> list[list[str]]:
        """读取表格单元格范围。"""
        from urllib.parse import quote

        encoded_fid = quote(file_id, safe="")
        encoded_sid = quote(sheet_id, safe="")
        data = self._get(
            f"/spreadsheet/v2/files/{encoded_fid}/sheets/{encoded_sid}/values",
            params={"range": cell_range},
        )
        return data.get("values", [])

    def _put_json(self, path: str, json_body: dict[str, Any]) -> dict[str, Any]:
        """PUT 请求，发送 JSON body。"""
        resp = self._session.put(self._url(path), json=json_body, timeout=30)
        resp.raise_for_status()
        return self._check_response(resp.json())

    def _request_raw_url(
        self,
        method: str,
        raw_url: str,
        json_body: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """直接使用完整 URL 发送请求，不经过 _url() 编码。

        用于 sheetbook API，其 URL 中的 ! 和 : 不能被二次编码。

        Args:
            method: HTTP 方法 ("PUT" / "POST")。
            raw_url: 完整 URL，已编码好。
            json_body: JSON 请求体。

        Returns:
            API 响应 data。
        """
        if method.upper() == "PUT":
            resp = self._session.put(raw_url, json=json_body, timeout=30)
        else:
            resp = self._session.post(raw_url, json=json_body, timeout=30)
        resp.raise_for_status()
        return self._check_response(resp.json())

    def _build_sheetbook_url(self, path: str) -> str:
        """构建 sheetbook API 的完整 URL。

        对 fileID 中的 $ 进行编码，但保留 range 中的 ! 和 : 不编码。

        Args:
            path: 路径，如 /sheetbook/v2/300000000$XXX/values/BB08J2!A1:C2

        Returns:
            完整 URL。
        """
        from urllib.parse import quote

        # 拆分路径，只对 fileID 部分编码
        parts = path.split("/")
        result_parts = []
        for part in parts:
            # 包含 $ 的部分是 fileID，需要编码
            if "$" in part:
                result_parts.append(quote(part, safe=""))
            else:
                result_parts.append(part)
        return f"{self._base_url}{'/'.join(result_parts)}"

    def write_sheet(
        self,
        file_id: str,
        sheet_id: str,
        cell_range: str,
        values: list[list[Any]],
    ) -> dict[str, Any]:
        """写入表格单元格范围。

        Args:
            file_id: 表格文件 ID。
            sheet_id: 工作表 ID。
            cell_range: 写入范围，如 "A1:C3"。
            values: 二维数组，行数×列数必须匹配 range 维度。

        Returns:
            更新后的范围信息。

        Raises:
            ValueError: values 不是二维数组。
            TencentDocsError: API 返回错误。
        """
        if not values or not isinstance(values, list) or not isinstance(values[0], list):
            raise ValueError("values 必须是非空二维数组，如 [[\"a\",\"b\"],[\"c\",\"d\"]]")

        url = self._build_sheetbook_url(
            f"/sheetbook/v2/{file_id}/values/{sheet_id}!{cell_range}"
        )
        return self._request_raw_url("PUT", url, json_body={"values": values})

    def clear_sheet(
        self, file_id: str, sheet_id: str, cell_range: str
    ) -> dict[str, Any]:
        """清空表格指定范围。

        Args:
            file_id: 表格文件 ID。
            sheet_id: 工作表 ID。
            cell_range: 清空范围，如 "A1:Z100"。

        Returns:
            清空后的范围信息。
        """
        url = self._build_sheetbook_url(
            f"/sheetbook/v2/{file_id}/values/{sheet_id}!{cell_range}:clear"
        )
        return self._request_raw_url("POST", url, json_body={})

    def _batch_update(
        self, file_id: str, operation: dict[str, Any]
    ) -> dict[str, Any]:
        """执行批量结构更新（添加/删除子表、行、列）。

        注意：腾讯文档的 batchUpdate 不使用 requests 数组包装，
        操作对象直接放在顶层 JSON body 中。

        Args:
            file_id: 表格文件 ID。
            operation: 操作对象，如 {"addSheet": {"properties": {"title": "新子表"}}}。

        Returns:
            批量更新结果。
        """
        url = self._build_sheetbook_url(
            f"/sheetbook/v2/{file_id}:batchUpdate"
        )
        return self._request_raw_url("POST", url, json_body=operation)

    def add_sheet(self, file_id: str, title: str) -> dict[str, Any]:
        """添加新的子表（工作表）。

        Args:
            file_id: 表格文件 ID。
            title: 新子表的标题。

        Returns:
            新增子表的信息（含 sheet_id）。
        """
        return self._batch_update(
            file_id,
            {"addSheet": {"properties": {"title": title}}},
        )

    def delete_rows(
        self, file_id: str, sheet_id: str, start_index: int, count: int = 1
    ) -> dict[str, Any]:
        """删除指定工作表的行。

        Args:
            file_id: 表格文件 ID。
            sheet_id: 工作表 ID。
            start_index: 起始行号（0-based，即第1行为0）。
            count: 删除行数，默认 1。

        Returns:
            删除结果。
        """
        return self._batch_update(
            file_id,
            {"deleteDimension": {
                "range": {
                    "sheetID": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": start_index,
                    "endIndex": start_index + count,
                }
            }},
        )

    def delete_columns(
        self, file_id: str, sheet_id: str, start_index: int, count: int = 1
    ) -> dict[str, Any]:
        """删除指定工作表的列。

        Args:
            file_id: 表格文件 ID。
            sheet_id: 工作表 ID。
            start_index: 起始列号（0-based，即A列为0）。
            count: 删除列数，默认 1。

        Returns:
            删除结果。
        """
        return self._batch_update(
            file_id,
            {"deleteDimension": {
                "range": {
                    "sheetID": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": start_index,
                    "endIndex": start_index + count,
                }
            }},
        )

    # ── 收集表 ────────────────────────────────────────────

    def generate_form_result(self, file_id: str) -> dict[str, Any]:
        """生成/获取收集表的结果表。"""
        from urllib.parse import quote

        encoded_id = quote(file_id, safe="")
        return self._post(f"/drive/v2/forms/{encoded_id}/result")

    def read_form(self, file_id: str) -> list[dict[str, Any]]:
        """读取收集表内容（自动：获取结果表 → 异步导出 → 下载解析）。

        策略：
        1. 先从元数据的 relativeFiles 获取已关联的结果表 ID（最可靠）。
        2. 若无关联结果表，尝试调用 generate_form_result 生成。
        3. 对结果表做异步导出 → 下载 → 解析 xlsx。

        Returns:
            字段列表 + 提交记录列表，格式:
            [{"sheet_name": ..., "fields": [...], "records": [[...], ...]}]
        """
        result_file_id = ""

        # 1. 优先从元数据获取已关联的结果表
        meta = self.read_metadata(file_id)
        rel = meta.get("relativeFiles", [])
        if rel:
            result_file_id = rel[0].get("fileID", "")

        # 2. 若无关联结果表，尝试生成
        if not result_file_id:
            try:
                result_info = self.generate_form_result(file_id)
                result_file_id = result_info.get("ID", "")
            except TencentDocsError:
                pass

        if not result_file_id:
            raise TencentDocsError(-1, "无法获取收集表结果表 ID，请确认收集表已有提交数据")

        # 2. 异步导出
        xlsx_bytes = self.async_export(result_file_id, export_type="xlsx")

        # 3. 解析 xlsx
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
        result: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            wb.close()

            if not rows:
                continue

            fields = rows[0]
            records = rows[1:]
            result.append({
                "sheet_name": sheet_name,
                "fields": fields,
                "records": records,
            })

        return result

    # ── 异步导出 ──────────────────────────────────────────

    def async_export(
        self,
        file_id: str,
        export_type: str = "xlsx",
        max_wait: int = 30,
        poll_interval: float = 2.0,
    ) -> bytes:
        """异步导出文件，自动轮询进度并下载。

        Args:
            file_id: 文件 ID。
            export_type: 导出格式 (xlsx / pdf)。
            max_wait: 最大等待秒数。
            poll_interval: 轮询间隔秒数。

        Returns:
            导出文件的二进制内容。

        Raises:
            TencentDocsError: 导出超时或失败。
        """
        from urllib.parse import quote

        encoded_id = quote(file_id, safe="")

        # Step 1: 发起导出
        export_data = self._post(
            f"/drive/v2/files/{encoded_id}/async-export",
            json_body={"exportType": export_type},
        )
        operation_id = export_data.get("operationID", "")
        if not operation_id:
            raise TencentDocsError(-1, "异步导出未返回 operationID")

        # Step 2: 轮询进度
        deadline = time.monotonic() + max_wait
        while time.monotonic() < deadline:
            time.sleep(poll_interval)
            progress_data = self._get(
                f"/drive/v2/files/{encoded_id}/export-progress",
                params={"operationID": operation_id},
            )
            progress = progress_data.get("progress", 0)
            if progress >= 100:
                download_url = progress_data.get("url", "")
                break
        else:
            raise TencentDocsError(-1, f"异步导出超时 ({max_wait}s)")

        if not download_url:
            raise TencentDocsError(-1, "导出完成但未返回下载链接")

        # Step 3: 下载
        resp = requests.get(download_url, timeout=60)
        resp.raise_for_status()
        return resp.content

    # ── 创建文件 ──────────────────────────────────────────

    def create_file(
        self,
        title: str,
        doc_type: str = "doc",
        folder_id: str | None = None,
    ) -> dict[str, Any]:
        """创建文件。

        Args:
            title: 文件标题。
            doc_type: 文件类型 (doc/sheet/slide/mind/flowchart/smartsheet/form)。
            folder_id: 目标文件夹 ID，None 为根目录。

        Returns:
            包含 ID 和 url 的字典。
        """
        data: dict[str, str] = {"title": title, "type": doc_type}
        if folder_id is not None:
            data["folderID"] = folder_id
        return self._post("/drive/v2/files", data=data)

    # ── 上传图片 ──────────────────────────────────────────

    def upload_image(self, image_path: str) -> dict[str, Any]:
        """上传图片（用于嵌入文档）。

        Args:
            image_path: 图片文件路径。

        Returns:
            上传结果。
        """
        with open(image_path, "rb") as f:
            resp = self._session.post(
                self._url("/resources/v2/images"),
                files={"file": f},
                timeout=60,
            )
        resp.raise_for_status()
        return self._check_response(resp.json())

    # ── 重命名 ────────────────────────────────────────────

    def rename(self, file_id: str, new_title: str) -> dict[str, Any]:
        """重命名文件。⚠️ 写操作，需先确认。"""
        from urllib.parse import quote

        encoded_id = quote(file_id, safe="")
        return self._patch(
            f"/drive/v2/files/{encoded_id}",
            data={"title": new_title},
        )

    # ── 移动 ──────────────────────────────────────────────

    def move(self, file_id: str, folder_id: str) -> dict[str, Any]:
        """移动文件到指定文件夹。⚠️ 写操作，需先确认。"""
        from urllib.parse import quote

        encoded_id = quote(file_id, safe="")
        return self._patch(
            f"/drive/v2/files/{encoded_id}/move",
            data={"folderID": folder_id},
        )

    # ── 复制 ──────────────────────────────────────────────

    def copy(self, file_id: str, title: str) -> dict[str, Any]:
        """复制文件。"""
        from urllib.parse import quote

        encoded_id = quote(file_id, safe="")
        return self._post(
            f"/drive/v2/files/{encoded_id}/copy",
            data={"title": title},
        )

    # ── 删除 ──────────────────────────────────────────────

    def delete(self, file_id: str) -> dict[str, Any]:
        """删除文件（移入回收站）。⚠️ 写操作，需先确认。"""
        from urllib.parse import quote

        encoded_id = quote(file_id, safe="")
        return self._delete(f"/drive/v2/files/{encoded_id}")
